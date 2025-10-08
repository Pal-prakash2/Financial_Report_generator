"""Excel workbook builder for XBRL parsing results."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Dict, Iterable, List, Mapping, MutableMapping, Sequence

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.validation_service import ValidationMessage
from app.services.xbrl_parser import AuditRecord, UnmappedFact, XBRLParseResult

STATEMENT_SHEETS = {
    "income_statement": "Income Statement",
    "balance_sheet": "Balance Sheet",
    "cash_flow": "Cash Flow",
}

HEADER_FILL = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
ISSUE_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


class ExcelGenerator:
    """Create Excel workbooks from parsed XBRL statement bundles."""

    def generate(self, parse_result: XBRLParseResult, validations: Sequence[ValidationMessage]) -> BytesIO:
        workbook = Workbook()
        # Remove the default sheet once we add our own content.
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        issues_lookup = self._build_issue_lookup(validations)
        for statement_key, sheet_name in STATEMENT_SHEETS.items():
            sheet = workbook.create_sheet(sheet_name)
            self._write_statement_sheet(sheet, parse_result.statement(statement_key), issues_lookup.get(statement_key, {}))

        self._write_audit_sheet(workbook, parse_result.audit_trail, validations, parse_result.metadata)
        self._write_unmapped_sheet(workbook, parse_result.unmapped_facts)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _build_issue_lookup(validations: Sequence[ValidationMessage]) -> Dict[str, Dict[str, Dict[str, ValidationMessage]]]:
        lookup: Dict[str, Dict[str, Dict[str, ValidationMessage]]] = {}
        for message in validations:
            lookup.setdefault(message.statement, {}).setdefault(message.field, {})[message.period] = message
        return lookup

    def _write_statement_sheet(self, sheet, data: Mapping[str, Mapping[str, object]], issues: Dict[str, Dict[str, ValidationMessage]]) -> None:
        sheet.append(["Metric"])
        periods = self._collect_periods(data)
        for idx, period in enumerate(periods, start=2):
            cell = sheet.cell(row=1, column=idx, value=period)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
        metric_cell = sheet.cell(row=1, column=1, value="Metric")
        metric_cell.font = Font(bold=True)
        metric_cell.fill = HEADER_FILL
        metric_cell.alignment = CENTER_ALIGN

        current_row = 2
        for field, period_values in sorted(data.items()):
            sheet.cell(row=current_row, column=1, value=field)
            for col_offset, period in enumerate(periods, start=2):
                value = period_values.get(period)
                cell = sheet.cell(row=current_row, column=col_offset, value=float(value) if value is not None else None)
                issue = issues.get(field, {}).get(period)
                if issue and not issue.passed:
                    cell.fill = ISSUE_FILL
                    if issue.message:
                        cell.comment = Comment(issue.message, "Validation")
                cell.number_format = "#,##0.00"
            current_row += 1

        sheet.freeze_panes = "B2"
        if periods:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(periods) + 1)}{current_row - 1}"
        sheet.column_dimensions["A"].width = 40
        for idx in range(2, len(periods) + 2):
            sheet.column_dimensions[get_column_letter(idx)].width = 20

    @staticmethod
    def _collect_periods(data: Mapping[str, Mapping[str, object]]) -> List[str]:
        seen: List[str] = []
        for period_values in data.values():
            for period in period_values.keys():
                if period not in seen:
                    seen.append(period)
        return seen

    def _write_audit_sheet(
        self,
        workbook: Workbook,
        audit_trail: Sequence[AuditRecord],
        validations: Sequence[ValidationMessage],
        metadata: Mapping[str, object],
    ) -> None:
        sheet = workbook.create_sheet("Audit Trail")
        headers = [
            "Concept",
            "Statement",
            "Field",
            "Period",
            "Value (INR)",
            "Unit",
            "Context Ref",
        ]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN

        for row_index, record in enumerate(audit_trail, start=2):
            sheet.cell(row=row_index, column=1, value=record.concept)
            sheet.cell(row=row_index, column=2, value=STATEMENT_SHEETS.get(record.statement, record.statement))
            sheet.cell(row=row_index, column=3, value=record.field)
            sheet.cell(row=row_index, column=4, value=record.period)
            sheet.cell(row=row_index, column=5, value=float(record.value))
            sheet.cell(row=row_index, column=6, value=record.unit)
            sheet.cell(row=row_index, column=7, value=record.context_ref)

        issue_start_row = len(audit_trail) + 3
        sheet.cell(row=issue_start_row, column=1, value="Validation Results").font = Font(bold=True)
        issue_headers = ["Statement", "Metric", "Period", "Status", "Details", "Difference"]
        for offset, header in enumerate(issue_headers):
            cell = sheet.cell(row=issue_start_row + 1, column=1 + offset, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL

        for idx, message in enumerate(validations, start=issue_start_row + 2):
            sheet.cell(row=idx, column=1, value=STATEMENT_SHEETS.get(message.statement, message.statement))
            sheet.cell(row=idx, column=2, value=message.field)
            sheet.cell(row=idx, column=3, value=message.period)
            sheet.cell(row=idx, column=4, value="PASS" if message.passed else "FAIL")
            sheet.cell(row=idx, column=5, value=message.message)
            sheet.cell(row=idx, column=6, value=float(message.difference))
            if not message.passed:
                for col in range(1, 7):
                    sheet.cell(row=idx, column=col).fill = ISSUE_FILL

        # Metadata block for quick reference
        meta_start = issue_start_row + len(validations) + 4
        sheet.cell(row=meta_start, column=1, value="Metadata").font = Font(bold=True)
        for index, (key, value) in enumerate(metadata.items(), start=meta_start + 1):
            key_cell = sheet.cell(row=index, column=1, value=str(key))
            value_cell = sheet.cell(row=index, column=2, value=str(value))
            if key == "source" and isinstance(value, str) and value:
                value_cell.hyperlink = value
                value_cell.font = Font(color="0563C1", underline="single")

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 25
        sheet.column_dimensions["E"].width = 20

    def _write_unmapped_sheet(self, workbook: Workbook, unmapped_facts: Sequence[UnmappedFact]) -> None:
        sheet = workbook.create_sheet("Unmapped Facts")
        headers = ["Concept", "Raw Tag", "Context", "Unit", "Raw Value"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN

        if not unmapped_facts:
            empty_cell = sheet.cell(row=2, column=1, value="All facts were mapped to known metrics.")
            empty_cell.font = Font(italic=True)
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            sheet.column_dimensions["A"].width = 40
            return

        for row_index, fact in enumerate(sorted(unmapped_facts, key=lambda item: (item.concept or "")), start=2):
            sheet.cell(row=row_index, column=1, value=fact.concept)
            sheet.cell(row=row_index, column=2, value=fact.raw_tag)
            sheet.cell(row=row_index, column=3, value=fact.context_ref)
            sheet.cell(row=row_index, column=4, value=fact.unit)
            sheet.cell(row=row_index, column=5, value=fact.raw_value)

        sheet.freeze_panes = "C2"
        sheet.auto_filter.ref = f"A1:E{len(unmapped_facts) + 1}"
        sheet.column_dimensions["A"].width = 45
        sheet.column_dimensions["B"].width = 60
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 50


__all__ = ["ExcelGenerator"]
