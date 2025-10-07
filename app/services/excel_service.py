from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.parsers import ParsedStatementBundle


class ExcelExporter:
    """Generate analyst-friendly Excel workbooks with audit trails."""

    def __init__(self, *, hyperlink_style: str = "Hyperlink") -> None:
        self.hyperlink_style = hyperlink_style

    def export(self, bundle: ParsedStatementBundle, output_path: str | Path) -> Path:
        workbook = Workbook()
        self._write_statement(workbook, "Balance Sheet", bundle.balance_sheet)
        self._write_statement(workbook, "Income Statement", bundle.income_statement)
        self._write_statement(workbook, "Cash Flow", bundle.cash_flow)

        metadata_sheet = workbook.create_sheet("Metadata")
        metadata_sheet["A1"] = "Source"
        metadata_sheet["B1"] = bundle.metadata.get("source", "")
        source_url = bundle.metadata.get("source")
        if source_url:
            cell = metadata_sheet["B1"]
            cell.hyperlink = source_url
            cell.style = self.hyperlink_style
        metadata_sheet["A2"] = "Financial Year"
        metadata_sheet["B2"] = bundle.metadata.get("financial_year")
        metadata_sheet.column_dimensions["A"].width = 20
        metadata_sheet.column_dimensions["B"].width = 60

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path

    def _write_statement(self, workbook: Workbook, sheet_name: str, data: dict[str, object]) -> None:
        sheet = workbook.create_sheet(sheet_name)
        sheet["A1"] = "Metric"
        sheet["B1"] = "Value (INR)"
        row = 2
        for metric, value in data.items():
            sheet[f"A{row}"] = metric
            sheet[f"B{row}"] = float(value)
            row += 1
        sheet.auto_filter.ref = f"A1:{get_column_letter(2)}{row - 1}"
        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 25

        for col in range(1, 3):
            sheet.cell(row=1, column=col).style = "Title"

        if "Sheet" in workbook.sheetnames:
            default_sheet = workbook["Sheet"]
            workbook.remove(default_sheet)
